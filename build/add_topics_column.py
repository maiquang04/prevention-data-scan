# -*- coding: utf-8 -*-
"""One-off: add the Topics column to the scan workbook.

A source used to belong to exactly one topic because of where it sat - the workbook
has a sheet per topic and a row lives on one of them. Nothing could be on two sheets
at once. The Topics column removes that limit.

Every row is pre-filled with its own sheet's topic, so running this changes nothing
on the site. Adding a second topic to a row is an editorial judgement, made by hand
afterwards.

    python build/add_topics_column.py

Safe to run twice: rows that already have a value are left alone. Close the workbook
in Excel first or openpyxl cannot save.
"""
import json
import os
import sys
from copy import copy

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DEFAULT_SRC = os.path.normpath(os.path.join(REPO, "..", "lit-scan", "lit-scan-draft.xlsx"))
TOPICS_FILE = os.path.join(REPO, "data", "topics.json")

HEADER_ROW = 2
COLUMN = "Topics"


def load_topics():
    with open(TOPICS_FILE, encoding="utf-8") as handle:
        return json.load(handle)["topics"]


def main(src=DEFAULT_SRC):
    topics = load_topics()
    by_sheet = {t["sheet"]: t for t in topics if t.get("sheet")}

    workbook = load_workbook(src)
    filled = skipped = 0

    for sheet in workbook:
        default = by_sheet.get(sheet.title)
        if default is None:
            print("No topic claims sheet %r - leaving its rows blank." % sheet.title,
                  file=sys.stderr)

        headers = {cell.value: cell.column for cell in sheet[HEADER_ROW] if cell.value}
        if COLUMN not in headers:
            col = sheet.max_column + 1
            source = sheet.cell(row=HEADER_ROW, column=1)
            cell = sheet.cell(row=HEADER_ROW, column=col, value=COLUMN)
            for attr in ("font", "fill", "border", "alignment"):
                setattr(cell, attr, copy(getattr(source, attr)))
            sheet.column_dimensions[cell.column_letter].width = 34
            headers[COLUMN] = col

        for row in sheet.iter_rows(min_row=HEADER_ROW + 1):
            if row[0].value is None:
                continue
            cell = sheet.cell(row=row[0].row, column=headers[COLUMN])
            if cell.value not in (None, ""):
                skipped += 1
                continue
            if default is None:
                continue
            cell.value = default["title"]
            # Keep Rocky's row highlighting unbroken across the new column.
            for attr in ("fill", "alignment", "border"):
                setattr(cell, attr, copy(getattr(row[0], attr)))
            filled += 1

    workbook.save(src)
    print("Filled %d row(s), left %d already-set row(s) alone." % (filled, skipped))
    print("Saved %s" % src)
    print("\nTo put a source under more than one topic, add to its Topics cell with a")
    print("semicolon, e.g.  Local government measurement; Health inequities")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC))
