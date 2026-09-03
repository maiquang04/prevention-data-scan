# -*- coding: utf-8 -*-
"""Export the literature and data scan workbook to JSON for the website.

The workbook is the source of truth. Thanh maintains it and Rocky annotates it by
highlighting rows, so nothing here is hand-edited - run this script instead:

    python build/export_data.py

Writes data/studies.json and data/meta.json. data/quadrants.json is hand-authored
and is never touched by this script.

Close the workbook in Excel before running, or openpyxl will read a stale copy.
"""
import argparse
import datetime
import json
import os
import re
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DEFAULT_SRC = os.path.normpath(os.path.join(REPO, "..", "lit-scan", "lit-scan-draft.xlsx"))
OUT_DIR = os.path.join(REPO, "data")
TOPICS_FILE = os.path.join(OUT_DIR, "topics.json")
INDICATORS_FILE = os.path.join(OUT_DIR, "indicators.json")
QUADRANTS_FILE = os.path.join(OUT_DIR, "quadrants.json")

HEADER_ROW = 2

# Workbook column heading -> JSON key. Anything not listed here is ignored, so a
# new column in the workbook will not silently leak onto a public website.
FIELDS = {
    "#": "num",
    "Reference": "reference",
    "Link": "link",
    "Source type": "sourceType",
    "Country/system": "country",
    "Task described": "task",
    "Metrics / indicators": "metrics",
    "Geographic level": "geoLevel",
    "Input data": "inputData",
    "Output": "output",
    "Data sources used": "dataSources",
    "Public data sources?": "accessNote",
    "Data sources link": "dataLinkRaw",
    "Scale": "scale",
    "Key attributes": "keyAttributes",
    "Access": "access",
    "Geo tags": "geoTagsRaw",
    "Topics": "topicsRaw",
    "Domain": "domainsRaw",
    "Indicators": "indicatorsRaw",
}

# Controlled vocabulary. Anything outside these lists is a data-entry mistake and
# the script stops rather than shipping a filter option nobody meant to create.
ACCESS_VALUES = [
    "Public",
    "Public (aggregate only)",
    "Restricted - application",
    "Not a dataset",
    "Not yet assessed",
]
GEO_TAGS = [
    "Address",
    "Small area",
    "LGA",
    "PHN",
    "HHS",
    "Remoteness",
    "State",
    "National",
    "International",
    "Individual",
    "Any",
]

# Which health issue a source speaks to. Not our vocabulary: these are the seven
# domains from the partner agency's own measures spreadsheet, kept in their wording
# and their order so a tag here means to them what it means here. Seven is few
# enough to stay in code; the 61 indicators underneath them are not, and belong in
# a data file if they are ever added.
#
# Blank is allowed and common. A source about how to build an index, or about
# hospital admissions, has no health-issue domain, and inventing one to fill the
# column would be worse than leaving it empty.
DOMAINS = [
    "Physical Activity",
    "Healthy Eating",
    "Implementation",
    "Social",
    "Equity",
    "Prosperity & Productivity",
    "Mental Wellbeing",
]

# Two derived facets. The workbook's own "Source type" and "Country/system" columns
# are prose and have around a dozen distinct values each, which makes a useless
# filter. These roll them into buckets people actually filter by. The original
# wording is still shown on the record.
SOURCE_GROUPS = [
    "Peer-reviewed",
    "Government report",
    "Published dataset",
    "International agency report",
    "Academic or institutional report",
]
REGIONS = ["Queensland", "Australia", "International"]

# First rule whose wording appears in the cell wins, so order matters: "Peer-reviewed,
# plus a government data tool" is peer-reviewed, not a government report.
#
# This used to end in a bare `return "Academic or institutional report"`, which meant
# any wording nobody had thought of - a blank cell included - was quietly filed as an
# academic report. The counts still added up and nothing flagged it. That is exactly
# the failure the Access and Geo tags columns are designed to prevent, so unmatched
# wording is now a data-entry error like any other.
SOURCE_GROUP_RULES = [
    ("Peer-reviewed", ("peer-reviewed", "peer reviewed", "journal article")),
    # Renamed from "Statistical agency release". Of the 15 rows it holds, two are not
    # statistical agencies at all: PHIDU is a university unit and Queensland Globe is a
    # mapping service. Both are still published datasets, which is what a reader wants
    # to know, so the bucket now says that instead of guessing at who published it.
    ("Published dataset", ("statistical agency", "statistician", "official statistics",
                           "statistics release", "survey data", "administrative data",
                           "census", "data release", "dataset", "spatial data")),
    ("Government report", ("government", "department of", "ministry", "public agency")),
    ("International agency report", ("international agency", "world health organization",
                                     "united nations", "oecd")),
    ("Academic or institutional report", ("academic", "institutional", "university",
                                          "think tank", "communique", "statement of principles",
                                          "collective")),
]

URL_RE = re.compile(r"https?://\S+")


def source_group(source_type):
    """The bucket this source is filtered under, or None if the wording matches no
    rule - the caller turns that into a data-entry error rather than a guess."""
    text = source_type.lower()
    for group, needles in SOURCE_GROUP_RULES:
        if any(needle in text for needle in needles):
            return group
    return None


def regions_of(country):
    """Australian and Queensland evidence is wanted first, so that is the split
    worth filtering on - not the fifteen different country strings.

    Returns None for a blank cell. It used to return International, which is a
    plausible-looking answer to a question nobody answered - an unfilled row would
    appear in the filters as overseas evidence."""
    text = country.lower().strip()
    if not text:
        return None
    tags = []
    if "queensland" in text:
        tags.append("Queensland")
    if "australia" in text:
        tags.append("Australia")
    if not tags or any(word in text for word in ("global", "countries", "oecd", "many")):
        tags.append("International")
    return tags


def publication_year(reference):
    """The year in the reference, for the newest/oldest sort on the browse page.

    Every reference that has a year writes it the same way, in brackets after the
    authors: "Masters R, ... (2017). Return on investment ...". So the bracketed form
    is the only one read. A bare four-digit number elsewhere in the string is usually
    part of a title or a survey name and would date the source wrongly.

    Returns None for the 20 rows with no year, which are the ongoing collections -
    the National Health Survey, PLIDA, the Social Health Atlas, AusPlay. They have no
    single publication year, and giving them one would be inventing it. The browse
    page sorts them to the bottom whichever direction is chosen."""
    match = re.search(r"\((\d{4})[a-z]?\)", reference)
    return int(match.group(1)) if match else None


def clean(value):
    """Collapse whitespace; treat blanks and None alike."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def priority_of(row):
    """Rocky marks relevance by filling cells: yellow relevant, red highly relevant."""
    for cell in row:
        fill = cell.fill
        if fill is not None and fill.fill_type == "solid":
            if getattr(fill.start_color, "rgb", None) == "FFFFFF00":
                return "relevant"
            if getattr(fill.start_color, "theme", None) == 5:
                return "highly-relevant"
    return "unmarked"


def parse_links(text):
    """Split a free-text cell into labelled links.

    Cells hold anything from a bare URL to 'Office for National Statistics:
    https://... Deprivation data: https://...'. Take each URL, and use whatever
    text sits in front of it as the label.
    """
    text = clean(text)
    if not text:
        return []
    links, cursor = [], 0
    for match in URL_RE.finditer(text):
        label = text[cursor:match.start()].strip(" -:;,")
        url = match.group(0).rstrip(".,;)")
        links.append({"label": label, "url": url})
        cursor = match.end()
    if not links:
        # No URL at all - the cell is a note, not a link.
        return [{"label": text, "url": ""}]
    return links


def app_number(sheet_title):
    """Only used to build a stable study id. It records which sheet a row lives on,
    which is not the same thing as which topics it belongs to."""
    match = re.search(r"App\s*(\d+)", sheet_title)
    if not match:
        raise SystemExit("Sheet name does not carry a number: " + sheet_title)
    return int(match.group(1))


def load_topics():
    """The taxonomy lives in data/topics.json so a new topic is a data change, not a
    code change. See that file's own comment for the shape."""
    with open(TOPICS_FILE, encoding="utf-8") as handle:
        topics = json.load(handle)["topics"]

    seen = set()
    for topic in topics:
        for field in ("id", "title", "heading"):
            if not topic.get(field):
                raise SystemExit("topics.json: an entry is missing '%s'" % field)
        if topic["id"] in seen:
            raise SystemExit("topics.json: duplicate id '%s'" % topic["id"])
        seen.add(topic["id"])
    return topics


def load_indicators():
    """The 61 measures come from the partner agency and change when they revise them,
    so they live in data/indicators.json rather than in a list here. The seven domains
    stay in code above because they are few and stable.

    Returns {name, domain} in file order. Order matters: it is the agency's own ranking
    and the pages show them in it. The domain travels with the name because the browse
    filter groups all 61 under their seven domain headings - sixty-one flat checkboxes
    would drown the six filter groups below them."""
    with open(INDICATORS_FILE, encoding="utf-8") as handle:
        entries = json.load(handle)["indicators"]

    out, seen = [], set()
    for entry in entries:
        name = (entry.get("name") or "").strip()
        if not name:
            raise SystemExit("indicators.json: an entry has no 'name'")
        if name in seen:
            raise SystemExit("indicators.json: duplicate indicator '%s'" % name)
        domain = (entry.get("domain") or "").strip()
        if domain and domain not in DOMAINS:
            raise SystemExit("indicators.json: '%s' has domain '%s', which is not one of "
                             "the seven" % (name, domain))
        seen.add(name)
        out.append({"name": name, "domain": domain})
    return out


def gridded_topic_ids():
    """Which topics have a hand-authored quadrant grid. Lets the pages stop hardcoding
    that Application 1 is the only one built."""
    try:
        with open(QUADRANTS_FILE, encoding="utf-8") as handle:
            return set(json.load(handle).get("topics", {}))
    except (OSError, ValueError):
        return set()


def export(src, out_dir, quiet=False):
    if not os.path.exists(src):
        raise SystemExit("Workbook not found: " + src)

    topics = load_topics()
    indicators_meta = load_indicators()
    indicator_names = [i["name"] for i in indicators_meta]
    topic_by_title = {t["title"]: t["id"] for t in topics}
    topic_by_sheet = {t["sheet"]: t["id"] for t in topics if t.get("sheet")}

    workbook = load_workbook(src, data_only=True)
    studies, problems = [], []

    for sheet in workbook:
        app = app_number(sheet.title)
        default_topic = topic_by_sheet.get(sheet.title)

        headers = {cell.value: idx for idx, cell in enumerate(sheet[HEADER_ROW]) if cell.value}
        missing = [h for h in FIELDS if h not in headers]
        if missing:
            problems.append("Sheet '%s' is missing columns: %s" % (sheet.title, ", ".join(missing)))
            continue

        for row in sheet.iter_rows(min_row=HEADER_ROW + 1):
            if row[0].value is None:
                continue

            record = {}
            for heading, key in FIELDS.items():
                record[key] = clean(row[headers[heading]].value)

            num = int(record.pop("num"))
            where = "App %d row %d" % (app, num)

            access = record.get("access", "")
            if access not in ACCESS_VALUES:
                problems.append("%s: Access is '%s', not one of the allowed values" % (where, access))

            tags = [t.strip() for t in record.pop("geoTagsRaw", "").split(";") if t.strip()]
            for tag in tags:
                if tag not in GEO_TAGS:
                    problems.append("%s: geo tag '%s' is not in the vocabulary" % (where, tag))
            if not tags:
                problems.append("%s: no geo tags" % where)

            # Same split and check as the geo tags, without the "must have at least
            # one" rule - plenty of sources genuinely have no health-issue domain.
            domains = [d.strip() for d in record.pop("domainsRaw", "").split(";") if d.strip()]
            for domain in domains:
                if domain not in DOMAINS:
                    problems.append("%s: domain '%s' is not in the vocabulary" % (where, domain))

            # Same again for the indicators, and blank is the usual answer here. A
            # source earns a tag only by naming or computing the measure; the agency
            # said they would fill in the rest, so gaps are the plan and not a lapse.
            indicators = [i.strip() for i in record.pop("indicatorsRaw", "").split(";") if i.strip()]
            for indicator in indicators:
                if indicator not in indicator_names:
                    problems.append("%s: indicator '%s' is not in indicators.json"
                                    % (where, indicator))

            # A blank cell means "whatever this sheet is about", so the column only
            # has to be filled in where a source genuinely spans more than one topic.
            named = [t.strip() for t in record.pop("topicsRaw", "").split(";") if t.strip()]
            topic_ids = []
            for name in named:
                if name not in topic_by_title:
                    problems.append("%s: topic '%s' is not in topics.json" % (where, name))
                elif topic_by_title[name] not in topic_ids:
                    topic_ids.append(topic_by_title[name])
            if not topic_ids:
                if default_topic:
                    topic_ids = [default_topic]
                else:
                    problems.append("%s: no topic, and no topic claims sheet '%s'"
                                    % (where, sheet.title))

            record["id"] = "app%d-%02d" % (app, num)
            record["app"] = app
            record["num"] = num
            record["topics"] = topic_ids
            record["priority"] = priority_of(row)
            record["geoTags"] = tags
            record["domains"] = domains
            record["indicators"] = indicators
            group = source_group(record.get("sourceType", ""))
            if group is None:
                problems.append("%s: source type '%s' matches no rule in "
                                "SOURCE_GROUP_RULES - add a rule or fix the cell"
                                % (where, record.get("sourceType", "")))
            record["sourceGroup"] = group or ""

            regions = regions_of(record.get("country", ""))
            if regions is None:
                problems.append("%s: no country or system given" % where)
            record["regions"] = regions or []
            record["dataLinks"] = parse_links(record.pop("dataLinkRaw", ""))

            record["year"] = publication_year(record["reference"])

            if not record["reference"]:
                problems.append("%s: no reference" % where)

            studies.append(record)

    if problems:
        print("Export stopped - fix the workbook first:", file=sys.stderr)
        for problem in problems:
            print("  - " + problem, file=sys.stderr)
        raise SystemExit(1)

    studies.sort(key=lambda s: (s["app"], s["num"]))

    counts = {}
    for study in studies:
        counts[study["access"]] = counts.get(study["access"], 0) + 1

    gridded = gridded_topic_ids()
    meta = {
        "generated": datetime.date.today().isoformat(),
        "source": os.path.basename(src),
        "total": len(studies),
        "topics": [
            {
                "id": t["id"],
                "title": t["title"],
                "heading": t["heading"],
                "question": t.get("question", ""),
                "blurb": t.get("blurb", ""),
                "count": sum(1 for s in studies if t["id"] in s["topics"]),
                "hasGrid": t["id"] in gridded,
                # The old ?app=1 links already sent out have to keep resolving.
                # Derived from the sheet name, the same way a study id is, so this
                # cannot drift out of step with it.
                "legacyApp": app_number(t["sheet"]) if t.get("sheet") else None,
            }
            for t in topics
        ],
        "byAccess": counts,
        "accessValues": ACCESS_VALUES,
        "geoTags": GEO_TAGS,
        "domains": DOMAINS,
        "indicators": indicators_meta,
        "sourceGroups": SOURCE_GROUPS,
        "regions": REGIONS,
        "withDownloadableData": sum(
            1 for s in studies if s["access"] in ("Public", "Public (aggregate only)")
        ),
    }

    os.makedirs(out_dir, exist_ok=True)
    for name, payload in (("studies.json", studies), ("meta.json", meta)):
        path = os.path.join(out_dir, name)
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, ensure_ascii=False)
        if not quiet:
            print("wrote %s (%d KB)" % (path, os.path.getsize(path) // 1024))

    if not quiet:
        print("\n%d studies: %s" % (
            len(studies),
            ", ".join("%s = %d" % (t["title"], t["count"]) for t in meta["topics"]),
        ))
        for value in ACCESS_VALUES:
            print("  %-26s %d" % (value, counts.get(value, 0)))
        dated = [s2["year"] for s2 in studies if s2["year"]]
        print()
        print("%d of %d sources carry a publication year (%d-%d); "
              "%d are ongoing collections with none"
              % (len(dated), len(studies), min(dated), max(dated),
                 len(studies) - len(dated)))

    # Re-stamp the asset links here too, so the one command people actually run
    # leaves the site in a publishable state.
    if out_dir == OUT_DIR:
        import stamp_assets
        if not quiet:
            print("")
        stamp_assets.stamp(quiet)

    return studies, meta


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--src", default=DEFAULT_SRC, help="path to lit-scan-draft.xlsx")
    parser.add_argument("--out", default=OUT_DIR, help="output directory for the JSON")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args()
    export(args.src, args.out, args.quiet)

