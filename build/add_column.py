# -*- coding: utf-8 -*-
"""Add an empty column to every sheet of the scan workbook.

    python build/add_column.py "Domain"
    python build/add_column.py "Indicators" --width 44

Generalised out of add_topics_column.py, which did the same thing for one hardcoded
column and also pre-filled it. Filling is left to whoever adds the column here,
because there is no sensible default for an arbitrary column.

The styling matters more than it looks. Rocky marks relevance by filling the row
(yellow relevant, red highly relevant) and export_data.py's priority_of() reads
those fills back off the cells. An unstyled column dropped into a highlighted row
would leave a white gap through it, and the marking is then only as reliable as
whichever cell happens to be checked first. So the new cells inherit the row's fill,
border and alignment even though they hold nothing yet.

Safe to run twice: a sheet that already has the column is left alone. Close the
workbook in Excel first or openpyxl cannot save.
"""
import argparse
import os
import sys
from copy import copy

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DEFAULT_SRC = os.path.normpath(os.path.join(REPO, "..", "lit-scan", "lit-scan-draft.xlsx"))

HEADER_ROW = 2


def add_column(src, column, width=34):
    workbook = load_workbook(src)
    added = existing = styled = 0

    for sheet in workbook:
        headers = {cell.value: cell.column for cell in sheet[HEADER_ROW] if cell.value}
        if column in headers:
            existing += 1
            print("  %-32s already has it" % sheet.title)
            continue

        col = sheet.max_column + 1
        heading = sheet.cell(row=HEADER_ROW, column=1)
        cell = sheet.cell(row=HEADER_ROW, column=col, value=column)
        for attr in ("font", "fill", "border", "alignment"):
            setattr(cell, attr, copy(getattr(heading, attr)))
        sheet.column_dimensions[cell.column_letter].width = width

        for row in sheet.iter_rows(min_row=HEADER_ROW + 1):
            if row[0].value is None:
                continue
            new = sheet.cell(row=row[0].row, column=col)
            for attr in ("fill", "alignment", "border"):
                setattr(new, attr, copy(getattr(row[0], attr)))
            styled += 1

        added += 1
        print("  %-32s added as column %s" % (sheet.title, cell.column_letter))

    if not added:
        print("\nNothing to do - every sheet already has %r." % column)
        return 0

    workbook.save(src)
    print("\nAdded %r to %d sheet(s), carried the row styling onto %d cell(s)."
          % (column, added, styled))
    if existing:
        print("Left %d sheet(s) that already had it alone." % existing)
    print("Saved %s" % src)
    print("\nThe column is empty. Fill it before adding it to FIELDS in export_data.py,")
    print("or the export will ship a column of blanks.")
    return 0


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("column", help="the column heading to add")
    parser.add_argument("--width", type=int, default=34, help="column width (default 34)")
    parser.add_argument("--src", default=DEFAULT_SRC, help="workbook to edit")
    args = parser.parse_args(argv)

    if not os.path.exists(args.src):
        raise SystemExit("Workbook not found: " + args.src)

    print("Adding %r to %s\n" % (args.column, args.src))
    return add_column(args.src, args.column, args.width)


if __name__ == "__main__":
    sys.exit(main())
